import os
import pandas as pd
import math
import xlsxwriter
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
df = pd.read_excel(r"D:\R&D_2023\sample1.xlsx")
#print(df)
v=df['v(t)']
T=df['time(sec)']
T=list(df['time(sec)'])
n=len(v)
deltaT=0.00002
v1=[]
w=(math.pi)*100
#for i in range(1,n-1):
    #vdash=((v[i+1]-v[i-1])/(0.00004))
    #v_dash.append(vdash)
#print(v_dash)
for j in range(1,n-1):
    #vdash = ((v[j + 1] - v[j - 1]) / (0.00004))
    #v_dash.append(vdash)
    vm=math.sqrt((v[j]**2)+((v[j+1]-v[j-1])/(w*0.00004))**2)
    v1.append(vm)
    if (vm>253 or vm<207):
        print("Fault detected")
        print(v1)  # it is vm
        print(T[j + 1])
        print("this is time at which fault occured")
        break
    else:
        continue


    #v1.append(vm)
#d1={}
#d1 = dict(zip(vm,T))
#print(d1)
#d1 = dict(map(lambda m,n : (m,n) , vm,T))
#print(d1)
#print("vm")

# print(v1) #it is vm

#print(T[j+1])
# print("this is time at which fault occured")
#vm.to_excel("sample1.xlsx",index=false)


# if os.path.isfile('sample1'):
#     workbook=openpyxl.load_workbook('sample1')
#     sheet=workbook['sheet1']
#     for row in dataframe_to_rows(df,header=False,index=False):
#         sheet.append(row+1)
#     workbook.save('sample1')
#     workbook.close()
result= [math.ceil(item) for item in vm]
workbook=xlsxwriter.Workbook('sample1.xlsx')
sheet=workbook.add_worksheet()

sheet.write('C1','vm')
row=1
column=3
for z in result:
    worksheet.write(row,column,z)
    row+=1

# df2 = pd.DataFrame(v1)
# df2.to_excel('sample1.xlsx', sheet_name='sheet1')
# print(df2)
# df3 = df[['a', 'c']]


# book = xlsxwriter.Book('sample1.xlsx')
# sheet = book.add_sheet()
#
# # Rows and columns are zero indexed.
# row = 1
# column = 4
#
# for item in v1:
#     # write operation perform
#     sheet.write(row, column, item)
#
#     # incrementing the value of row by one with each iterations.
#     row += 1
#
# book.close()

#outWorkbook = xlsxwriter.Workbook('D:\R&D_2023\sample1.xlsx')
#outSheet = outWorkbook.add_worksheet()

#outSheet.write("D1", "Names")

#for item in range(len(v1)):
 #   outSheet.write(item+1, 4, v1[item])


#outWorkbook.close()